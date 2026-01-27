"""Export analysis results to Excel, PDF, and CSV."""

import csv
import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

logger = logging.getLogger(__name__)


def export_to_excel(story_title: str, analysis: dict) -> bytes:
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Abuse Cases sheet
    ws = wb.active
    ws.title = "Abuse Cases"
    headers = ["ID", "Threat", "Actor", "Description", "Impact", "Likelihood", "Attack Vector", "STRIDE"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    for row_idx, ac in enumerate(analysis.get("abuse_cases", []), 2):
        ws.cell(row=row_idx, column=1, value=ac.get("id", "")).border = border
        ws.cell(row=row_idx, column=2, value=ac.get("threat", "")).border = border
        ws.cell(row=row_idx, column=3, value=ac.get("actor", "")).border = border
        ws.cell(row=row_idx, column=4, value=ac.get("description", "")).border = border
        ws.cell(row=row_idx, column=5, value=ac.get("impact", "")).border = border
        ws.cell(row=row_idx, column=6, value=ac.get("likelihood", "")).border = border
        ws.cell(row=row_idx, column=7, value=ac.get("attack_vector", "")).border = border
        ws.cell(row=row_idx, column=8, value=ac.get("stride_category", "")).border = border
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20

    # Security Requirements sheet
    ws2 = wb.create_sheet("Security Requirements")
    headers2 = ["ID", "Requirement", "Priority", "Category", "Details"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    for row_idx, req in enumerate(analysis.get("security_requirements", []), 2):
        ws2.cell(row=row_idx, column=1, value=req.get("id", "")).border = border
        ws2.cell(row=row_idx, column=2, value=req.get("text", "")).border = border
        ws2.cell(row=row_idx, column=3, value=req.get("priority", "")).border = border
        ws2.cell(row=row_idx, column=4, value=req.get("category", "")).border = border
        ws2.cell(row=row_idx, column=5, value=req.get("details", "")).border = border
    for col in range(1, 6):
        ws2.column_dimensions[chr(64 + col)].width = 25

    # STRIDE Threats sheet
    ws3 = wb.create_sheet("STRIDE Threats")
    headers3 = ["Category", "Threat", "Description", "Risk Level"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    for row_idx, st in enumerate(analysis.get("stride_threats", []), 2):
        ws3.cell(row=row_idx, column=1, value=st.get("category", "")).border = border
        ws3.cell(row=row_idx, column=2, value=st.get("threat", "")).border = border
        ws3.cell(row=row_idx, column=3, value=st.get("description", "")).border = border
        ws3.cell(row=row_idx, column=4, value=st.get("risk_level", "")).border = border

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_csv(analysis: dict) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow(["Section", "ID", "Title/Threat", "Description", "Severity/Priority", "Category"])

    for ac in analysis.get("abuse_cases", []):
        writer.writerow(["Abuse Case", ac.get("id", ""), ac.get("threat", ""), ac.get("description", ""), ac.get("impact", ""), ac.get("stride_category", "")])

    for req in analysis.get("security_requirements", []):
        writer.writerow(["Requirement", req.get("id", ""), req.get("text", ""), req.get("details", ""), req.get("priority", ""), req.get("category", "")])

    for st in analysis.get("stride_threats", []):
        writer.writerow(["STRIDE Threat", "", st.get("threat", ""), st.get("description", ""), st.get("risk_level", ""), st.get("category", "")])

    return buf.getvalue().encode("utf-8")


def export_to_pdf(story_title: str, analysis: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("CustomTitle", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#6B21A8"))
    heading_style = ParagraphStyle("CustomHeading", parent=styles["Heading2"], textColor=colors.HexColor("#6B21A8"))

    elements.append(Paragraph("Security Analysis Report", title_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"User Story: {story_title}", styles["Heading3"]))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Paragraph(f"Risk Score: {analysis.get('risk_score', 0)}/100", styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Abuse Cases
    elements.append(Paragraph("Abuse Cases", heading_style))
    elements.append(Spacer(1, 8))
    ac_data = [["ID", "Threat", "Impact", "Likelihood"]]
    for ac in analysis.get("abuse_cases", []):
        ac_data.append([ac.get("id", ""), ac.get("threat", "")[:50], ac.get("impact", ""), ac.get("likelihood", "")])
    if len(ac_data) > 1:
        t = Table(ac_data, colWidths=[60, 250, 70, 70])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6B21A8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F3FF")]),
        ]))
        elements.append(t)
    elements.append(Spacer(1, 20))

    # Security Requirements
    elements.append(Paragraph("Security Requirements", heading_style))
    elements.append(Spacer(1, 8))
    req_data = [["ID", "Requirement", "Priority", "Category"]]
    for req in analysis.get("security_requirements", []):
        req_data.append([req.get("id", ""), req.get("text", "")[:60], req.get("priority", ""), req.get("category", "")[:20]])
    if len(req_data) > 1:
        t = Table(req_data, colWidths=[50, 230, 70, 100])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6B21A8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F3FF")]),
        ]))
        elements.append(t)

    doc.build(elements)
    return buf.getvalue()
